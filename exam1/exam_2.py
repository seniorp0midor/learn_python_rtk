import csv
import fastavro
import openpyxl
from openpyxl.utils import get_column_letter
import csv
import pandas as pd

# конвертим файл из xlsx в csv
wb = openpyxl.load_workbook('ozon.xlsx')
sheet = wb.active
headers = [cell.value for cell in sheet[1]]
data = []
for row in sheet.iter_rows(values_only=True):
    data.append(row)
with open('output.csv', 'w', newline='', errors='ignore', encoding='utf-8') as file:
    writer = csv.writer(file, delimiter=';')
    writer.writerow(headers)
    writer.writerows(data)


# читаем CSV и получаем уникальные бренды
def get_unique_brands_from_csv(file_path):
    unique_brands = set()
    with open(file_path, 'r', encoding='utf-8') as file:
        reader = csv.DictReader(file, delimiter=';')
        for row in reader:
            brand_name = row.get('Бренд')
            if brand_name:
                unique_brands.add(brand_name)
    return unique_brands


def write_to_avro_file(data, schema, file_path):
    records = []
    for i, brand_name in enumerate(data):
        records.append({"number": i + 1, "brand_name": brand_name})

    with open(file_path, 'wb') as fo:
        fastavro.writer(fo, schema, records)

csv_file_path = 'output.csv'

unique_brands = get_unique_brands_from_csv(csv_file_path)

# Схема
schema = {
    "type": "record",
    "name": "Brand",
    "fields": [
        {"name": "number", "type": "int"},
        {"name": "brand_name", "type": "string"}
    ]
}

avro_file_path = 'ozon.avro'

write_to_avro_file(unique_brands, schema, avro_file_path)
print(f"Уникальные бренды сохранены в файл {avro_file_path}.")


def read_avro_file(file_path):
    with open(file_path, 'rb') as fo:
        reader = fastavro.reader(fo)
        records = [record for record in reader]
    return records

records = read_avro_file(avro_file_path)
for record in records:
    print(record)
